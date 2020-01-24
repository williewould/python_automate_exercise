# -*- coding: utf-8 -*-
'''
Created on 2020/01/24
@Version:    1.0
@author :    zhaowu
@Ref    :    利用openpyxl模块来操作excel文件
'''

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 基本表格文件的属性操作
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
print(wb.sheetnames)
sheet3 = wb['Sheet3']
print(type(sheet3))
print(sheet3.title)
sheet = wb.active
print(type(sheet))
sheet['A1']
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s ' % (c.coordinate, c.value))
print(sheet['C1'].value)

# 使用 行列 数字坐标索引列表
sheet.cell(row=1, column=2)
sheet.cell(row=1, column=2).value
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
sheet.max_row
sheet.max_column
# 列表数字坐标和字母坐标之间的转换
print(get_column_letter(1)+' '+get_column_letter(2)+' '+get_column_letter(27) +
      ' '+get_column_letter(900)+' '+get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

# 元组方式 索引行列值
tuple(sheet['A1':'C3'])
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

# 列表方式 索引行类值
list(sheet.columns)[1]
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
for cellObj in list(sheet.rows)[1]:
    print(cellObj.value)
