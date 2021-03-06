# -*- coding: utf-8 -*-
'''
Created on 2020/01/25
@Version:    1.0
@author :    zhaowu
@Ref    :    book
@Funcion:    练习用openpyxl创建Excel文档
'''


import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
# 第一个 从0开始 新建sheet
wb.create_sheet(index=0, title='Fist Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
del wb['Middle Sheet']
print(wb.sheetnames)
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!'
print(sheet['A1'].value)
# 修改字体
fontObj = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj
sheet['A1'] = 'Bold Times New Roman'
fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'
wb.save('styles.xlsx')
# %% 在excel中写入excel公式
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2['A1'] = 200
sheet2['A2'] = 300
sheet2['A3'] = '=SUM(A1:A2)'
wb2.save('writeFormula.xlsx')

# %% 在excel修改cell的尺寸
wb3 = openpyxl.Workbook()
sheet3 = wb3.active
sheet3['A1'] = 'Tall row'
sheet3['B2'] = 'Wide column'
sheet3.row_dimensions[1].height = 70
sheet3.column_dimensions['B'].width = 20
wb3.save('dimensions.xlsx')

# %% 合并单元格
wb5 = openpyxl.Workbook()
sheet5 = wb5.active
sheet5.merge_cells('A1:D3')
sheet5['A1'] = 'Twelve cells merged together.'
sheet5.merge_cells('C5:D5')
sheet5['C5'] = 'Tow merged cells.'
wb5.save('merged.xlsx')

# %% freeze panes 冻结窗格
wb6 = openpyxl.load_workbook('produceSales.xlsx')
sheet6 = wb6.active
sheet6.freeze_panes = 'A2'
wb6.save('freezeExample.xlsx')
# %% chart 图表
wb7 = openpyxl.Workbook()
sheet7 = wb7.active
for i in range(1, 11):
    sheet7['A'+str(i)] = i
# 划定取用数据的范围refObj,+序列名字，构成一个series
refObj = openpyxl.chart.Reference(
    sheet7, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
# 其他类型 .LineChart(), .ScatterChart(), PieChart()
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
# 利用series生成chart
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb7.save('sampleChart.xlsx')
